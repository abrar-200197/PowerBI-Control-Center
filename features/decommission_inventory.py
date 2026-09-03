"""
Inventory of decommissioned reports from SharePoint archive tree.

Layout (same as archive upload):
  {SHAREPOINT_DECOMM_FOLDER_PATH}/
    <batch folder>/
      <WorkspaceName>/
        [<PBI Folder>/]
          Report.pbix | Report.rdl

We do not claim live Power BI metadata — only SharePoint file facts:
  workspace, optional folder, report name, type, size, dates, webUrl, batch.

Workspace folders with zero archived reports are still listed (reportCount=0)
so tracking placeholders appear in workspace count / filters.
"""

from __future__ import annotations

import logging
import time
from datetime import datetime, timezone
from pathlib import PurePosixPath
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# Process cache (Graph walks are chatty)
_CACHE: Dict[str, Any] = {"ts": 0.0, "payload": None}
_CACHE_TTL_SEC = 300  # 5 minutes


def _parse_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        t = s.replace("Z", "+00:00")
        return datetime.fromisoformat(t)
    except Exception:
        return None


def _fmt_dt(s: Optional[str]) -> Optional[str]:
    dt = _parse_dt(s)
    if not dt:
        return s
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _strip_ext(name: str) -> Tuple[str, str]:
    p = PurePosixPath(name)
    ext = (p.suffix or "").lower()
    stem = p.stem if ext else name
    return stem, ext.lstrip(".")


def _is_report_file(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith(".pbix") or n.endswith(".rdl")


def _row_from_file(item: Dict[str, Any], base: str) -> Optional[Dict[str, Any]]:
    """
    Map one driveItem under decomm root → inventory row.
    relativePath is full path from drive root (set by list_files_recursive).
    """
    name = item.get("name") or ""
    if not _is_report_file(name):
        return None

    rel = (item.get("relativePath") or "").replace("\\", "/").strip("/")
    base_n = base.strip("/")
    if base_n and rel.startswith(base_n + "/"):
        under = rel[len(base_n) + 1 :]
    elif base_n and rel == base_n:
        return None
    else:
        under = rel

    parts = [p for p in under.split("/") if p]
    # Expect: batch / workspace / [folder…] / file
    if len(parts) < 3:
        # batch/file or workspace missing — skip incomplete paths
        if len(parts) == 2:
            batch, fname = parts
            workspace, folder = "Unknown", None
            if not _is_report_file(fname):
                return None
        else:
            return None
    else:
        batch = parts[0]
        workspace = parts[1]
        fname = parts[-1]
        mid = parts[2:-1]
        folder = " / ".join(mid) if mid else None
        if not _is_report_file(fname):
            return None

    report_name, ext = _strip_ext(fname)
    created = item.get("createdDateTime") or item.get("fileSystemInfo", {}).get("createdDateTime")
    modified = item.get("lastModifiedDateTime")
    # Decommissioned date = upload/create on SharePoint (archive time)
    decomm_raw = created or modified
    size = int(item.get("size") or 0)

    return {
        "reportName": report_name,
        "fileName": fname,
        "fileType": ext.upper() if ext else "FILE",
        "workspaceName": workspace,
        "folderName": folder,
        "batchFolder": batch,
        "decommissionedAt": decomm_raw,
        "decommissionedAtDisplay": _fmt_dt(decomm_raw),
        "lastModifiedAt": modified,
        "lastModifiedAtDisplay": _fmt_dt(modified),
        "sizeBytes": size,
        "sizeDisplay": _size_label(size),
        "webUrl": item.get("webUrl") or "",
        "sharePointPath": rel,
    }


def _size_label(n: int) -> str:
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    return f"{n / (1024 * 1024):.1f} MB"


def build_decommission_inventory(*, force_refresh: bool = False) -> Dict[str, Any]:
    """
    Walk SharePoint Report Decommission Activity tree → flat report rows + workspace groups.
    Cached in-process for _CACHE_TTL_SEC.
    """
    global _CACHE
    now = time.time()
    if (
        not force_refresh
        and _CACHE.get("payload")
        and (now - float(_CACHE.get("ts") or 0)) < _CACHE_TTL_SEC
    ):
        return dict(_CACHE["payload"])

    from catalog_service import catalog_config as cfg
    from catalog_service.metadata_lib.sharepoint_client import SharePointClient

    base = (getattr(cfg, "SHAREPOINT_DECOMM_FOLDER_PATH", None) or "").strip("/")
    if not base:
        return {
            "success": False,
            "error": "SHAREPOINT_DECOMM_FOLDER_PATH is not configured",
            "rows": [],
            "workspaces": [],
        }

    try:
        sp = SharePointClient()
        sp.resolve_site_and_drive()
        # Confirm base exists
        try:
            batches = sp.list_folders(base)
        except Exception as ex:
            return {
                "success": False,
                "error": f"Cannot list decommission root '{base}': {ex}",
                "basePath": base,
                "rows": [],
                "workspaces": [],
            }

        files = sp.list_files_recursive(base, max_depth=10)
        rows: List[Dict[str, Any]] = []
        try:
            from catalog_service.thin_packs import is_excluded_report_name
        except Exception:
            def is_excluded_report_name(name):  # type: ignore
                n = (name or "").strip().casefold()
                return n in {
                    "usage metrics report",
                    "report usage metrics report",
                    "dashboard usage metrics report",
                }
        for it in files:
            row = _row_from_file(it, base)
            if not row:
                continue
            # Hide platform usage metrics archives (not business decomm content)
            if is_excluded_report_name(row.get("reportName")):
                continue
            rows.append(row)

        # Newest first
        def _sort_key(r: Dict[str, Any]):
            return r.get("decommissionedAt") or r.get("lastModifiedAt") or ""

        rows.sort(key=_sort_key, reverse=True)

        # Group by workspace (from report files)
        by_ws: Dict[str, List[Dict[str, Any]]] = {}
        for r in rows:
            wn = r.get("workspaceName") or "Unknown"
            by_ws.setdefault(wn, []).append(r)

        # Also include empty workspace folders under each batch (tracking placeholders).
        # Layout: base / <batch> / <WorkspaceName> / ...
        # Does not change report rows or totalReports — only workspace list/count.
        empty_ws = 0
        for b in batches:
            bname = (b.get("name") or "").strip()
            if not bname:
                continue
            batch_path = f"{base}/{bname}" if base else bname
            try:
                ws_folders = sp.list_folders(batch_path)
            except Exception as ex:
                logger.warning(
                    "decommission: list workspace folders failed under %s: %s",
                    batch_path,
                    ex,
                )
                continue
            for wf in ws_folders:
                wn = (wf.get("name") or "").strip()
                if not wn:
                    continue
                if wn not in by_ws:
                    by_ws[wn] = []
                    empty_ws += 1

        workspaces = []
        for wn, rs in sorted(by_ws.items(), key=lambda x: x[0].lower()):
            workspaces.append({
                "workspaceName": wn,
                "reportCount": len(rs),
                "reports": rs,
                "isEmpty": len(rs) == 0,
            })

        payload = {
            "success": True,
            "basePath": base,
            "batchFolders": sorted({b.get("name") for b in batches if b.get("name")}),
            "batchCount": len(batches),
            "totalReports": len(rows),
            "workspaceCount": len(workspaces),
            "emptyWorkspaceCount": empty_ws,
            "rows": rows,
            "workspaces": workspaces,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "cacheTtlSec": _CACHE_TTL_SEC,
            "source": "sharepoint",
            "note": (
                "Inventory is derived from SharePoint archive files only "
                "(not live Power BI metadata). Decommissioned date = file upload/create time. "
                "Workspace folders with no archived reports are included (0 reports)."
            ),
        }
        _CACHE = {"ts": now, "payload": payload}
        return dict(payload)
    except Exception as ex:
        logger.exception("decommission inventory failed")
        return {
            "success": False,
            "error": str(ex),
            "basePath": base,
            "rows": [],
            "workspaces": [],
        }


def _dataset_feed_folder() -> str:
    """
    Stable SharePoint folder for the Power BI programme dataset source files.
    Default: sibling of archive batches — …/Report Decommission Activity/_dataset_feed
    """
    import os
    from catalog_service import catalog_config as cfg

    override = (os.getenv("DECOMM_DATASET_FEED_FOLDER") or "").strip().strip("/")
    if override:
        return override
    base = (getattr(cfg, "SHAREPOINT_DECOMM_FOLDER_PATH", None) or "").strip("/")
    if not base:
        return "_dataset_feed"
    return f"{base}/_dataset_feed"


def _excel_bytes_from_rows(rows: List[Dict[str, Any]], *, generated_at: str) -> bytes:
    """Stable flat sheet for Power BI Get Data → Excel (SharePoint)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "DecommissionedReports"

    headers = [
        "Report",
        "FileName",
        "Workspace",
        "Folder",
        "Type",
        "Decommissioned",
        "DecommissionedAt",
        "Batch",
        "Size",
        "SizeBytes",
        "SizeGB",
        "FileURL",
        "SharePointPath",
        "GeneratedAtUTC",
    ]
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        size_b = int(r.get("sizeBytes") or 0)
        size_gb = round(size_b / (1024 ** 3), 6) if size_b else 0.0
        ws.append([
            r.get("reportName") or "",
            r.get("fileName") or "",
            r.get("workspaceName") or "",
            r.get("folderName") or "",
            r.get("fileType") or "",
            r.get("decommissionedAtDisplay") or "",
            r.get("decommissionedAt") or "",
            r.get("batchFolder") or "",
            r.get("sizeDisplay") or "",
            size_b,
            size_gb,
            r.get("webUrl") or "",
            r.get("sharePointPath") or "",
            generated_at,
        ])

    ws.auto_filter.ref = f"A1:N1"
    ws.freeze_panes = "A2"
    for letter, w in {
        "A": 36, "B": 36, "C": 28, "D": 22, "E": 10, "F": 22, "G": 24,
        "H": 36, "I": 12, "J": 14, "K": 12, "L": 40, "M": 48, "N": 22,
    }.items():
        ws.column_dimensions[letter].width = w

    # Meta sheet for refresh diagnostics in Power BI
    meta = wb.create_sheet("Meta")
    meta.append(["Key", "Value"])
    meta.append(["GeneratedAtUTC", generated_at])
    meta.append(["RowCount", len(rows)])
    meta.append(["Source", "Power BI Control Center decommission inventory"])
    meta.append([
        "Note",
        "Overwrite target for scheduled dataset refresh. Do not rename this workbook.",
    ])
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes_from_rows(rows: List[Dict[str, Any]], *, generated_at: str) -> bytes:
    import csv
    import io

    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow([
        "Report", "FileName", "Workspace", "Folder", "Type",
        "Decommissioned", "DecommissionedAt", "Batch", "Size", "SizeBytes",
        "SizeGB", "FileURL", "SharePointPath", "GeneratedAtUTC",
    ])
    for r in rows:
        size_b = int(r.get("sizeBytes") or 0)
        size_gb = round(size_b / (1024 ** 3), 6) if size_b else 0.0
        w.writerow([
            r.get("reportName") or "",
            r.get("fileName") or "",
            r.get("workspaceName") or "",
            r.get("folderName") or "",
            r.get("fileType") or "",
            r.get("decommissionedAtDisplay") or "",
            r.get("decommissionedAt") or "",
            r.get("batchFolder") or "",
            r.get("sizeDisplay") or "",
            size_b,
            size_gb,
            r.get("webUrl") or "",
            r.get("sharePointPath") or "",
            generated_at,
        ])
    return buf.getvalue().encode("utf-8-sig")


def publish_decommission_dataset_feed(
    *,
    force_inventory_refresh: bool = True,
    inventory: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Build flat inventory files and overwrite a fixed SharePoint path so a
    Power BI import dataset can refresh on a schedule (true 'live' overview).

    Files (replace in place):
      {DECOMM_DATASET_FEED_FOLDER}/Decommissioned_Inventory_Latest.xlsx
      {DECOMM_DATASET_FEED_FOLDER}/Decommissioned_Inventory_Latest.csv

    Does not modify archive batch folders or the Detail UI API contract.
    """
    import os
    import tempfile
    from pathlib import Path

    if inventory is None:
        inventory = build_decommission_inventory(force_refresh=force_inventory_refresh)
    if not inventory.get("success"):
        return {
            "success": False,
            "error": inventory.get("error") or "Inventory build failed",
            "stage": "inventory",
        }

    rows = list(inventory.get("rows") or [])
    generated_at = (
        inventory.get("generatedAt")
        or datetime.now(timezone.utc).isoformat()
    )
    folder = _dataset_feed_folder()
    xlsx_name = (
        os.getenv("DECOMM_DATASET_FEED_XLSX_NAME")
        or "Decommissioned_Inventory_Latest.xlsx"
    ).strip() or "Decommissioned_Inventory_Latest.xlsx"
    csv_name = (
        os.getenv("DECOMM_DATASET_FEED_CSV_NAME")
        or "Decommissioned_Inventory_Latest.csv"
    ).strip() or "Decommissioned_Inventory_Latest.csv"

    try:
        from catalog_service.metadata_lib.sharepoint_client import SharePointClient

        xlsx_bytes = _excel_bytes_from_rows(rows, generated_at=generated_at)
        csv_bytes = _csv_bytes_from_rows(rows, generated_at=generated_at)

        sp = SharePointClient()
        sp.resolve_site_and_drive()
        sp.ensure_folder(folder)

        uploaded = []
        with tempfile.TemporaryDirectory(prefix="decomm_feed_") as td:
            tdir = Path(td)
            xlsx_path = tdir / xlsx_name
            csv_path = tdir / csv_name
            xlsx_path.write_bytes(xlsx_bytes)
            csv_path.write_bytes(csv_bytes)

            for local, remote_name in ((xlsx_path, xlsx_name), (csv_path, csv_name)):
                remote = f"{folder}/{remote_name}"
                item = sp.upload_file(local, remote)
                uploaded.append({
                    "name": remote_name,
                    "remotePath": remote,
                    "webUrl": item.get("webUrl") or "",
                    "size": item.get("size") or local.stat().st_size,
                    "id": item.get("id"),
                })
                logger.info("decomm dataset feed uploaded %s", remote)

        return {
            "success": True,
            "folder": folder,
            "rowCount": len(rows),
            "generatedAt": generated_at,
            "files": uploaded,
            "xlsxPath": f"{folder}/{xlsx_name}",
            "csvPath": f"{folder}/{csv_name}",
            "xlsxUrl": next((f.get("webUrl") for f in uploaded if f["name"] == xlsx_name), ""),
            "csvUrl": next((f.get("webUrl") for f in uploaded if f["name"] == csv_name), ""),
            "note": (
                "Point the Power BI dataset at Decommissioned_Inventory_Latest.xlsx "
                "(or .csv) under this SharePoint folder, then schedule refresh."
            ),
        }
    except Exception as ex:
        logger.exception("decomm dataset feed publish failed")
        return {
            "success": False,
            "error": str(ex),
            "stage": "upload",
            "folder": folder,
            "rowCount": len(rows),
        }


def _pbi_sp_token() -> Optional[str]:
    """Service-principal token for Power BI API (Dataset.ReadWrite.All / workspace access)."""
    try:
        import msal
        import os

        tid = (os.getenv("TENANT_ID") or "").strip()
        cid = (os.getenv("CLIENT_ID") or "").strip()
        sec = (os.getenv("CLIENT_SECRET") or "").strip()
        if not (tid and cid and sec):
            return None
        app = msal.ConfidentialClientApplication(
            cid,
            authority=f"https://login.microsoftonline.com/{tid}",
            client_credential=sec,
        )
        result = app.acquire_token_for_client(
            scopes=["https://analysis.windows.net/powerbi/api/.default"]
        )
        return result.get("access_token") if isinstance(result, dict) else None
    except Exception as ex:
        logger.warning("decomm SP token failed: %s", ex)
        return None


def _short_pbi_error(text: str, limit: int = 280) -> str:
    """Pull a readable message from Power BI / Graph error JSON if present."""
    import json as _json

    raw = (text or "").strip()
    if not raw:
        return ""
    try:
        data = _json.loads(raw)
        err = data.get("error") if isinstance(data, dict) else None
        if isinstance(err, dict):
            msg = err.get("message") or err.get("code") or ""
            pbi = err.get("pbi.error") or err.get("pbi") or {}
            if isinstance(pbi, dict):
                details = pbi.get("details") or []
                extra = []
                for d in details if isinstance(details, list) else []:
                    if isinstance(d, dict):
                        v = d.get("message") or d.get("detail") or d.get("value")
                        if v:
                            extra.append(str(v))
                if extra:
                    msg = (msg + " — " if msg else "") + "; ".join(extra[:3])
            if msg:
                raw = str(msg)
    except Exception:
        pass
    raw = " ".join(raw.split())
    return raw if len(raw) <= limit else raw[: limit - 1] + "…"


def trigger_decommission_dataset_refresh(
    *,
    access_token: str = "",
    workspace_id: Optional[str] = None,
    dataset_id: Optional[str] = None,
    allow_service_principal: bool = True,
) -> Dict[str, Any]:
    """
    POST refreshes on the programme dataset (after feed publish).

    Tries:
      1) Caller access_token (usually signed-in user SSO)
      2) App service principal (TENANT_ID/CLIENT_ID/CLIENT_SECRET) if user token fails
         or lacks rights — SP must be Member/Contributor/Admin on the workspace.

    workspace_id/dataset_id from args or DECOMM_DASHBOARD_* env.
    """
    import os
    import re
    import requests

    workspace_id = (
        workspace_id
        or os.getenv("DECOMM_DASHBOARD_WORKSPACE_ID")
        or os.getenv("DECOMM_WORKSPACE_ID")
        or ""
    ).strip()
    dataset_id = (
        dataset_id
        or os.getenv("DECOMM_DASHBOARD_DATASET_ID")
        or os.getenv("DECOMM_DATASET_ID")
        or ""
    ).strip()
    guid_re = re.compile(
        r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
    )
    if dataset_id and not guid_re.match(dataset_id):
        return {
            "success": False,
            "error": (
                f"DECOMM_DASHBOARD_DATASET_ID looks invalid ({dataset_id!r}). "
                "Use the semantic model GUID from Power BI → Dataset settings → URL."
            ),
            "stage": "config",
            "skipped": False,
        }
    if workspace_id and not guid_re.match(workspace_id):
        return {
            "success": False,
            "error": (
                f"DECOMM_DASHBOARD_WORKSPACE_ID looks invalid ({workspace_id!r}). "
                "Use the workspace GUID from the app.powerbi.com/groups/{{id}}/ URL."
            ),
            "stage": "config",
        }
    if not dataset_id:
        return {
            "success": False,
            "error": (
                "Dataset id unknown. Feed was published; set DECOMM_DASHBOARD_DATASET_ID "
                "or open Overview once so the report/dataset can be discovered."
            ),
            "stage": "config",
            "skipped": True,
        }

    tokens: List[Tuple[str, str]] = []
    if (access_token or "").strip():
        tokens.append(("user", access_token.strip()))
    if allow_service_principal:
        sp = _pbi_sp_token()
        if sp and not any(t == sp for _, t in tokens):
            tokens.append(("service_principal", sp))

    if not tokens:
        return {
            "success": False,
            "error": (
                "No Power BI token available (user session empty and service principal "
                "TENANT_ID/CLIENT_ID/CLIENT_SECRET missing)."
            ),
            "stage": "auth",
        }

    # Empty JSON object body is widely accepted; bare empty / notifyOption as fallbacks
    bodies: List[Any] = [{}, None, {"notifyOption": "NoNotification"}]
    attempts: List[Dict[str, Any]] = []
    last_err = ""
    last_status = None
    last_url = ""
    last_auth = ""

    try:
        for auth_label, token in tokens:
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }
            urls = []
            if workspace_id:
                urls.append(
                    f"https://api.powerbi.com/v1.0/myorg/groups/{workspace_id}"
                    f"/datasets/{dataset_id}/refreshes"
                )
            urls.append(
                f"https://api.powerbi.com/v1.0/myorg/datasets/{dataset_id}/refreshes"
            )

            for url in urls:
                for body in bodies:
                    last_url = url
                    last_auth = auth_label
                    if body is None:
                        resp = requests.post(url, headers=headers, timeout=60)
                    else:
                        resp = requests.post(url, headers=headers, json=body, timeout=60)
                    last_status = resp.status_code
                    body_txt = (resp.text or "")[:500]
                    attempts.append({
                        "auth": auth_label,
                        "url": url.split("/myorg")[-1],
                        "httpStatus": resp.status_code,
                        "bodyMode": (
                            "empty_json" if body == {}
                            else ("none" if body is None else "notifyOption")
                        ),
                        "errorSnippet": _short_pbi_error(body_txt, 120) if not resp.ok else "",
                    })
                    if resp.status_code in (200, 202):
                        return {
                            "success": True,
                            "httpStatus": resp.status_code,
                            "workspaceId": workspace_id or None,
                            "datasetId": dataset_id,
                            "message": "Dataset refresh accepted",
                            "url": url,
                            "auth": auth_label,
                            "attempts": attempts[-6:],
                        }
                    last_err = _short_pbi_error(body_txt) or body_txt or f"HTTP {resp.status_code}"
                    # 404 / 400 → try next body/url; 401/403 → try next auth mode
                    if resp.status_code in (401, 403):
                        break
                if last_status in (401, 403):
                    break  # next auth token
                # 404 on workspace path → still try tenant path with same auth
            # continue to next auth if all URLs failed

        hint = ""
        if last_status in (401, 403):
            hint = (
                " Need Contributor (or higher) on the workspace/semantic model for the "
                "signed-in user, OR add the app service principal as Member/Contributor "
                "on that workspace (Dataset.ReadWrite.All alone is not enough)."
            )
        elif last_status == 404:
            hint = (
                " Dataset id or workspace may be wrong — confirm DECOMM_DASHBOARD_DATASET_ID "
                "is the semantic model GUID (not the report id)."
            )
        elif last_status == 400 and last_err:
            if "another refresh" in last_err.lower() or "in progress" in last_err.lower():
                hint = " A refresh is already running — wait and retry, or use Reload after it finishes."
            elif "not refreshable" in last_err.lower() or "premium" in last_err.lower():
                hint = " Model may not allow API refresh (mode/capacity)."
        elif last_status == 429 or (last_err and "capacity" in last_err.lower()):
            hint = " Refresh limit or capacity — try again later in Power BI service."

        return {
            "success": False,
            "httpStatus": last_status,
            "error": (last_err or f"HTTP {last_status}") + hint,
            "workspaceId": workspace_id or None,
            "datasetId": dataset_id,
            "stage": "refresh",
            "url": last_url,
            "auth": last_auth or None,
            "attempts": attempts[-8:],
        }
    except Exception as ex:
        logger.exception("decomm dataset refresh failed")
        return {
            "success": False,
            "error": str(ex),
            "stage": "refresh",
            "datasetId": dataset_id,
            "attempts": attempts[-8:],
        }


def wait_decommission_dataset_refresh(
    *,
    access_token: str,
    workspace_id: Optional[str] = None,
    dataset_id: Optional[str] = None,
    max_wait_sec: int = 90,
    poll_sec: float = 3.0,
) -> Dict[str, Any]:
    """
    Poll latest dataset refresh until Completed/Failed or timeout.
    Used after Sync data so Overview embed reloads with fresher data when possible.
    """
    import os
    import time
    import requests

    workspace_id = (
        workspace_id
        or os.getenv("DECOMM_DASHBOARD_WORKSPACE_ID")
        or os.getenv("DECOMM_WORKSPACE_ID")
        or ""
    ).strip()
    dataset_id = (
        dataset_id
        or os.getenv("DECOMM_DASHBOARD_DATASET_ID")
        or os.getenv("DECOMM_DATASET_ID")
        or ""
    ).strip()
    if not access_token or not dataset_id:
        return {
            "success": False,
            "skipped": True,
            "error": "token or dataset_id missing",
            "stage": "wait",
        }

    headers = {"Authorization": f"Bearer {access_token}"}
    if workspace_id:
        hist_url = (
            f"https://api.powerbi.com/v1.0/myorg/groups/{workspace_id}"
            f"/datasets/{dataset_id}/refreshes?$top=1"
        )
    else:
        hist_url = (
            f"https://api.powerbi.com/v1.0/myorg/datasets/{dataset_id}/refreshes?$top=1"
        )

    deadline = time.time() + max(15, int(max_wait_sec))
    last_status = ""
    last_row: Dict[str, Any] = {}
    try:
        while time.time() < deadline:
            resp = requests.get(hist_url, headers=headers, timeout=30)
            if resp.status_code == 404 and workspace_id:
                # Fallback to tenant-scoped history
                hist_url = (
                    f"https://api.powerbi.com/v1.0/myorg/datasets/"
                    f"{dataset_id}/refreshes?$top=1"
                )
                resp = requests.get(hist_url, headers=headers, timeout=30)
            if not resp.ok:
                return {
                    "success": False,
                    "error": f"refresh history HTTP {resp.status_code}: {(resp.text or '')[:200]}",
                    "stage": "wait",
                    "datasetId": dataset_id,
                    "workspaceId": workspace_id or None,
                }
            values = (resp.json() or {}).get("value") or []
            last_row = values[0] if values else {}
            last_status = str(last_row.get("status") or "").strip()
            # Unknown / Completed / Failed / Disabled / Cancelled
            low = last_status.lower()
            if low in ("completed", "failed", "disabled", "cancelled"):
                ok = low == "completed"
                return {
                    "success": ok,
                    "status": last_status,
                    "endTime": last_row.get("endTime"),
                    "startTime": last_row.get("startTime"),
                    "refreshType": last_row.get("refreshType"),
                    "serviceExceptionJson": last_row.get("serviceExceptionJson"),
                    "datasetId": dataset_id,
                    "workspaceId": workspace_id or None,
                    "stage": "wait",
                    "message": (
                        "Dataset refresh completed"
                        if ok
                        else f"Dataset refresh ended: {last_status or 'unknown'}"
                    ),
                }
            time.sleep(max(1.0, float(poll_sec)))

        return {
            "success": False,
            "timedOut": True,
            "status": last_status or "Unknown",
            "datasetId": dataset_id,
            "workspaceId": workspace_id or None,
            "stage": "wait",
            "message": (
                "Refresh still running after wait — Overview will reload; "
                "data may catch up shortly in the service."
            ),
            "lastRow": {
                "status": last_row.get("status"),
                "startTime": last_row.get("startTime"),
            },
        }
    except Exception as ex:
        logger.exception("decomm refresh wait failed")
        return {
            "success": False,
            "error": str(ex),
            "stage": "wait",
            "datasetId": dataset_id,
        }
